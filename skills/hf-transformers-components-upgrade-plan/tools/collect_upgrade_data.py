#!/usr/bin/env python3
"""
Collect upgrade data for mindone.transformers version comparison

This tool collects raw data about differences between two transformers versions
and outputs an Excel file for manual analysis.

Usage:
    python collect_upgrade_data.py --target v5.0.0
    python collect_upgrade_data.py --source v4.57.1 --target v5.0.0

Output:
    upgrade_data_{source}_to_{target}.xlsx - Excel file with raw data for analysis
"""

import os
import subprocess
import sys
import argparse
import re
from pathlib import Path


def parse_args():
    """Parse command line arguments"""
    parser = argparse.ArgumentParser(
        description='Collect upgrade data for mindone.transformers'
    )
    parser.add_argument(
        '-s', '--source',
        type=str,
        default=None,
        help='Source transformers version (e.g., v4.57.1). If not specified, auto-detected from mindone.'
    )
    parser.add_argument(
        '-t', '--target',
        type=str,
        required=True,
        help='Target transformers version (e.g., v5.0.0). Required.'
    )
    parser.add_argument(
        '--mindone-path',
        type=str,
        default='mindone',
        help='Path to mindone repository (default: mindone)'
    )
    parser.add_argument(
        '--tf-path',
        type=str,
        default='transformers',
        help='Path to transformers repository (default: transformers)'
    )
    return parser.parse_args()


def detect_mindone_version(mindone_path):
    """Detect mindone version from __init__.py"""
    init_file = os.path.join(mindone_path, 'mindone/transformers/__init__.py')
    if not os.path.exists(init_file):
        raise FileNotFoundError(f"Cannot find {init_file}")

    with open(init_file, 'r', encoding='utf-8') as f:
        content = f.read()

    match = re.search(r'__version__\s*=\s*["\']([^"\']+)["\']', content)
    if match:
        version = match.group(1)
        if not version.startswith('v'):
            version = 'v' + version
        return version
    else:
        raise ValueError(f"Cannot detect version from {init_file}")


def get_transformers_files(tag, repo_path='transformers'):
    """Get list of files in transformers at specific tag"""
    result = subprocess.run(
        ['git', '-C', repo_path, 'ls-tree', '-r', '--name-only', tag, '--', 'src/transformers/'],
        capture_output=True, text=True
    )
    if result.returncode != 0:
        raise RuntimeError(f"Failed to get files for tag {tag}: {result.stderr}")

    files = []
    for line in result.stdout.strip().split('\n'):
        if line and '/models/' not in line:
            path = line.replace('src/transformers/', '')
            files.append(path)
    return sorted(files)


def get_mindone_files(mindone_path='mindone/mindone/transformers'):
    """Get list of files in mindone transformers"""
    files = []
    for root, dirs, filenames in os.walk(mindone_path):
        dirs[:] = [d for d in dirs if d not in ['models', '__pycache__']]
        for filename in filenames:
            full_path = os.path.join(root, filename)
            full_path = full_path.replace('\\', '/')
            if '/models/' in full_path or '/__pycache__/' in full_path:
                continue
            if 'mindone/mindone/transformers/' in full_path:
                relative_path = full_path.split('mindone/mindone/transformers/')[1]
                files.append(relative_path)
    return sorted(files)


def get_git_diff(tag1, tag2, repo_path='transformers'):
    """Get git diff between two tags"""
    result = subprocess.run(
        ['git', '-C', repo_path, 'diff', '--name-status', tag1, tag2, '--', 'src/transformers/'],
        capture_output=True, text=True
    )
    if result.returncode != 0:
        raise RuntimeError(f"Failed to get diff between {tag1} and {tag2}: {result.stderr}")

    changes = []
    for line in result.stdout.strip().split('\n'):
        if not line.strip() or '/models/' in line:
            continue
        parts = line.split('\t')
        status = parts[0]

        change = {'status': status}
        if status == 'A':
            change['path'] = parts[1].replace('src/transformers/', '')
        elif status == 'D':
            change['path'] = parts[1].replace('src/transformers/', '')
        elif status == 'M':
            change['path'] = parts[1].replace('src/transformers/', '')
        elif status.startswith('R'):
            similarity = status[1:]
            change['similarity'] = similarity
            change['old_path'] = parts[1].replace('src/transformers/', '')
            change['new_path'] = parts[2].replace('src/transformers/', '')

        changes.append(change)

    return changes


def create_excel_output(source_version, target_version, tf_source_files, tf_target_files,
                        mindone_files, changes, output_name):
    """Create Excel file with raw data for manual analysis"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    except ImportError:
        print("Error: openpyxl is required. Install with: pip install openpyxl")
        return 1

    wb = Workbook()
    ws = wb.active
    ws.title = "Upgrade Plan"

    # Define colors
    colors = {
        'header': '4472C4',
        'modify': 'FFEB9C',      # Yellow
        'delete': 'FFC7CE',      # Red
        'rename': 'BDD7EE',      # Blue
        'add_rec': 'C6EFCE',     # Dark Green
        'add_opt': 'E2EFDA',     # Light Green
        'add_ign': 'D9D9D9',     # Gray
        'title': 'D9E1F2',
        'info': 'F2F2F2'
    }

    # Get file sets for analysis
    tf_source_set = set(tf_source_files)
    mindone_set = set(mindone_files)

    # Categorize changes
    rows = []
    additions_count = 0
    deletions_count = 0
    modifications_count = 0
    renames_count = 0

    for change in changes:
        status = change['status']

        if status == 'A':
            # Addition - check if true new feature or intentionally omitted
            path = change['path']
            folder = os.path.dirname(path) or '(root)'
            filename = os.path.basename(path)
            # Mark as new feature if not in source transformers
            is_new_feature = path not in tf_source_set
            rows.append({
                'folder': folder,
                'filename': filename,
                'path': path,
                'action': 'Add',
                'is_new_feature': is_new_feature,
                'note': f"New file in {target_version}. {'True new feature' if is_new_feature else 'Check if intentionally omitted in mindone'}"
            })
            additions_count += 1

        elif status == 'D':
            # Deletion - only include if file exists in mindone
            path = change['path']
            in_mindone = path in mindone_set
            if in_mindone:  # Only include if mindone has this file
                folder = os.path.dirname(path) or '(root)'
                filename = os.path.basename(path)
                rows.append({
                    'folder': folder,
                    'filename': filename,
                    'path': path,
                    'action': 'Delete',
                    'in_mindone': in_mindone,
                    'note': f"Deleted in {target_version}. Present in mindone - should delete"
                })
                deletions_count += 1

        elif status == 'M':
            # Modification - only include if file exists in mindone
            path = change['path']
            in_mindone = path in mindone_set
            if in_mindone:  # Only include if mindone has this file
                folder = os.path.dirname(path) or '(root)'
                filename = os.path.basename(path)
                rows.append({
                    'folder': folder,
                    'filename': filename,
                    'path': path,
                    'action': 'Modify',
                    'in_mindone': in_mindone,
                    'note': f"Modified in {target_version}. Present in mindone - should update"
                })
                modifications_count += 1

        elif status.startswith('R'):
            # Rename - only include if old file exists in mindone
            old_path = change['old_path']
            new_path = change['new_path']
            old_in_mindone = old_path in mindone_set
            new_in_mindone = new_path in mindone_set

            if old_in_mindone:  # Only include if mindone has the old file
                rows.append({
                    'folder': os.path.dirname(old_path) or '(root)',
                    'filename': os.path.basename(old_path),
                    'path': old_path,
                    'action': 'Rename',
                    'new_path': new_path,
                    'old_in_mindone': old_in_mindone,
                    'new_in_mindone': new_in_mindone,
                    'note': f"Renamed to {new_path}. Old file in mindone - rename needed"
                })
                renames_count += 1

    # Sort rows by folder then action
    action_order = {'Modify': 1, 'Delete': 2, 'Rename': 3, 'Add': 4}
    rows.sort(key=lambda x: (x['folder'], action_order.get(x['action'], 5), x['filename']))

    # Write header section (8 rows)
    headers = [
        f"MindONE Upgrade Plan: {source_version} -> {target_version}",
        f"Statistics: {additions_count} additions, {deletions_count} deletions, {modifications_count} modifications, {renames_count} renames",
        "Rule: Files in source transformers but not in mindone (same version) are intentionally omitted - DO NOT add them during upgrade",
        "Priority Guide: High=Core API/training/generation, Medium=Integrations/pipelines/utils, Low=CLI/platform-specific",
        "Transformers Repo Changes: (To be analyzed based on actual changes - check additions for new features, deletions for removed components)",
        "",
        "Instructions: Fill in Priority (High/Medium/Low) and update Note with your analysis. For Additions marked as 'Check if intentionally omitted', verify against same-version comparison.",
        ""
    ]

    for i, header in enumerate(headers, 1):
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=6)
        cell = ws.cell(row=i, column=1, value=header)
        cell.font = Font(bold=True if i <= 5 else False, size=12 if i == 1 else 10)
        cell.fill = PatternFill(start_color=colors['title'] if i <= 5 else colors['info'], end_color=colors['title'] if i <= 5 else colors['info'], fill_type='solid')
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # Write column headers
    header_row = 9
    column_headers = ['Folder', 'Filename', 'Full Path', 'Action', 'Priority', 'Note']
    for col, header in enumerate(column_headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color=colors['header'], end_color=colors['header'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Write data rows
    for row_idx, row_data in enumerate(rows, start=header_row + 1):
        ws.cell(row=row_idx, column=1, value=row_data['folder'])
        ws.cell(row=row_idx, column=2, value=row_data['filename'])
        ws.cell(row=row_idx, column=3, value=f"transformers/{row_data['path']}")
        ws.cell(row=row_idx, column=4, value=row_data['action'])
        ws.cell(row=row_idx, column=5, value='')  # Priority - to be filled manually
        ws.cell(row=row_idx, column=6, value=row_data['note'])

        # Apply color coding based on action
        action = row_data['action']
        if action == 'Modify':
            fill_color = colors['modify']
        elif action == 'Delete':
            fill_color = colors['delete']
        elif action == 'Rename':
            fill_color = colors['rename']
        elif action == 'Add':
            # For additions, use different colors based on whether it's a true new feature
            if row_data.get('is_new_feature', False):
                fill_color = colors['add_rec']  # Dark green for true new features
            else:
                fill_color = colors['add_opt']  # Light green for others (need to check)
        else:
            fill_color = colors['add_ign']

        for col in range(1, 7):
            ws.cell(row=row_idx, column=col).fill = PatternFill(
                start_color=fill_color, end_color=fill_color, fill_type='solid'
            )

    # Adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 60

    # Freeze header
    ws.freeze_panes = 'A10'

    # Save workbook
    wb.save(output_name)
    return 0


def main():
    args = parse_args()

    # Determine source version
    if args.source:
        source_version = args.source
        print(f"Using specified source version: {source_version}")
    else:
        print("Auto-detecting mindone version...")
        source_version = detect_mindone_version(args.mindone_path)
        print(f"Detected mindone version: {source_version}")

    target_version = args.target
    print(f"Target version: {target_version}")

    print("\n" + "=" * 80)
    print(f"Collecting Upgrade Data")
    print(f"{source_version} -> {target_version}")
    print("=" * 80)
    print()

    # Step 1: Get file lists
    print("Step 1: Getting file lists...")
    try:
        tf_source_files = get_transformers_files(source_version, args.tf_path)
        tf_target_files = get_transformers_files(target_version, args.tf_path)
        mindone_files = get_mindone_files(os.path.join(args.mindone_path, 'mindone/transformers'))
    except Exception as e:
        print(f"Error: {e}")
        return 1

    print(f"  Transformers {source_version} files: {len(tf_source_files)}")
    print(f"  Transformers {target_version} files: {len(tf_target_files)}")
    print(f"  MindONE current files: {len(mindone_files)}")

    # Step 2: Analyze same-version diff
    print("\nStep 2: Analyzing same-version differences...")
    tf_source_set = set(tf_source_files)
    mindone_set = set(mindone_files)
    only_in_tf = sorted(tf_source_set - mindone_set)
    common = sorted(tf_source_set & mindone_set)
    print(f"  Files only in transformers: {len(only_in_tf)}")
    print(f"  Common files: {len(common)}")

    # Step 3: Get git diff
    print("\nStep 3: Getting version diff...")
    try:
        changes = get_git_diff(source_version, target_version, args.tf_path)
    except Exception as e:
        print(f"Error: {e}")
        return 1

    print(f"  Total changes: {len(changes)}")

    # Step 4: Create Excel output
    print("\nStep 4: Creating Excel output...")
    output_name = f'upgrade_data_{source_version}_to_{target_version}.xlsx'

    # Handle file conflicts
    counter = 1
    base_name = f'upgrade_data_{source_version}_to_{target_version}'
    while os.path.exists(output_name):
        output_name = f'{base_name}_{counter}.xlsx'
        counter += 1

    result = create_excel_output(
        source_version, target_version,
        tf_source_files, tf_target_files,
        mindone_files, changes, output_name
    )

    if result == 0:
        print(f"\n{'=' * 80}")
        print(f"Successfully generated: {output_name}")
        print(f"{'=' * 80}")
        print("\nNext steps (see SKILL.md Step 4-6):")
        print("1. Understand Excel structure and rules (Step 4)")
        print(f"   - Files only in transformers (same version): {len(only_in_tf)}")
        print("   - Light green Add rows in this list should be marked Ignorable")
        print("2. Fill in the Excel document (Step 5)")
        print("   - Row 5: Write transformers repo change summary")
        print("   - Column E: Fill Priority (High/Medium/Low) for EVERY row")
        print("   - Column D: ONLY update Add rows to Recommended/Optional/Ignorable")
        print("              DO NOT change Modify/Delete/Rename actions!")
        print("   - Column F: Add detailed analysis in Note column")
        print("3. Final review (Step 6)")
        print("   - Verify all cells filled, check colors, save document")

    return result


if __name__ == '__main__':
    sys.exit(main())
